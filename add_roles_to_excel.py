#!/usr/bin/env python3
"""
Add Roles, Thresholds, and Findings Tags sheets to Excel documentation
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_roles_findings_sheets():
    """Add detailed role and findings sheets to workbook"""
    wb = load_workbook("FLEX_Comprehensive_Network_Documentation.xlsx")

    # Sheet: Participant Roles
    print("👥 Creating Participant Roles sheet...")
    ws = wb.create_sheet("Participant Roles", 1)

    ws['A1'] = "NETWORK PARTICIPANT ROLES & RELATIONSHIPS"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:F1')

    row = 3
    ws[f'A{row}'] = "ROLE"
    ws[f'B{row}'] = "DEFINITION"
    ws[f'C{row}'] = "POSITION IN FLOW"
    ws[f'D{row}'] = "KEY CHARACTERISTIC"
    ws[f'E{row}'] = "DATABASE TABLE"
    ws[f'F{row}'] = "RISK INDICATOR"

    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)

    row += 1
    roles_data = [
        ("SENDER", "Original SOL source funds intermediaries", "Layer 1: Money Origin", "Distributes to many addresses", "funder_incoming_transfers (sender_address)", "Fund distribution width"),
        ("FUNDER", "Intermediary wallet connecting sources to creators", "Layer 2: Bridge", "Receives from senders, sends to creators", "creator_funders, funder_incoming_transfers", "Creator count served"),
        ("CREATOR", "Token launcher receiving funder support", "Layer 3: Final Recipient", "Launches token after receiving SOL", "creator_funders, creator_outgoing_transfers", "Funding pattern analysis"),
    ]

    for role, definition, position, characteristic, table, indicator in roles_data:
        ws[f'A{row}'] = role
        ws[f'B{row}'] = definition
        ws[f'C{row}'] = position
        ws[f'D{row}'] = characteristic
        ws[f'E{row}'] = table
        ws[f'F{row}'] = indicator
        ws[f'A{row}'].font = Font(bold=True, color="0563C1")
        row += 1

    # Column sizing
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 25

    # Sheet: SOL Threshold & Filtering
    print("🔍 Creating SOL Threshold sheet...")
    ws = wb.create_sheet("SOL Threshold & Filtering", 2)

    ws['A1'] = "MINIMUM_SOL THRESHOLD & TRANSFER FILTERING"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:D1')

    row = 3
    ws[f'A{row}'] = "THRESHOLD VALUE"
    ws[f'B{row}'] = "0.001 SOL"
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "USD EQUIVALENT"
    ws[f'B{row}'] = "~$0.15 USD (at 150 USDC/SOL)"
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "CODE LOCATION"
    ws[f'B{row}'] = "funder_incoming_extractor.py:51"
    ws[f'A{row}'].font = Font(bold=True)

    row += 2
    ws[f'A{row}'] = "WHY FILTER?"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    filters = [
        ("Dust Transfers", "Network spam, test transactions, minimal amounts"),
        ("Fee Precision", "Small system fees or error corrections"),
        ("Noise Reduction", "Reduces false positives in pattern detection"),
        ("Performance", "Excludes millions of micro-transfers"),
        ("Data Quality", "Focuses on meaningful funding flows only"),
    ]

    for reason, explanation in filters:
        ws[f'A{row}'] = reason
        ws[f'B{row}'] = explanation
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 1
    ws[f'A{row}'] = "IMPACT ON ANALYSIS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    impacts = [
        ("Recorded Transfers", "Only >= 0.001 SOL transfers saved"),
        ("Database Size", "Filters out 30-40% of micro-transactions"),
        ("Network Analysis", "Focuses on meaningful flows"),
        ("Self-Funding Detection", "Works on actual meaningful amounts"),
        ("False Positives", "Significant reduction"),
    ]

    for item, impact in impacts:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = impact
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Sheet: Findings Tags Reference
    print("🏷️  Creating Findings Tags sheet...")
    ws = wb.create_sheet("Findings Tags Reference", 3)

    ws['A1'] = "COMPLETE FINDINGS TAGS REFERENCE"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:G1')

    row = 3
    headers = ['Tag', 'Emoji', 'Risk Level', 'Detection Trigger', 'What It Means', 'Example', 'Action']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row += 1
    tags_data = [
        ("SELF-FUNDING", "🚩", "CRITICAL", ">50% self-created funders", "Creator owns multiple wallets, distributes to self", "24 of 28 funders are creator's own wallets", "Investigate pump-and-dump"),
        ("CREATOR_FUNDING_CHAIN", "⚠️", "HIGH", "Funder funded by another creator", "Creators funding each other's launches", "Funder X was funded by Creator C", "Check coordinated network"),
        ("DISTRIBUTION_PATTERN", "⚠️", "HIGH", "Recipients > 5× Funders", "Suspicious recipient redistribution", "10 funders → 85 recipients", "Monitor next tokens"),
        ("COORDINATED_FUNDERS", "🔗", "HIGH", "Shared funders with other creators", "Multiple creators share funders", "Funder X funds A, B, and C", "Map entire network"),
        ("NETWORK_MEMBER", "⚠️", "MEDIUM", "Part of detected network", "Creator in funder network cluster", "Member of FUNDERS_14 network", "Check network stats"),
        ("AUTOMATION_DETECTED", "🤖", "MEDIUM", "Funded by automation bot", "Funding from RapidLaunch/Axiom", "Creator funded by Axiom bot", "Check bot patterns"),
        ("INSTITUTIONAL_BACKED", "💱", "LOW", "Funded by known CEX", "Funding from exchange account", "Creator funded by Coinbase", "Reduces suspicion"),
        ("CLEAN", "✅", "NONE", "No suspicious patterns", "Organic, legitimate funding", "Normal funding pattern", "Standard monitoring"),
    ]

    for tag, emoji, risk, trigger, meaning, example, action in tags_data:
        ws[f'A{row}'] = tag
        ws[f'B{row}'] = emoji
        ws[f'C{row}'] = risk
        ws[f'D{row}'] = trigger
        ws[f'E{row}'] = meaning
        ws[f'F{row}'] = example
        ws[f'G{row}'] = action

        # Color risk cells
        risk_color = {
            'CRITICAL': 'FFCCCC',  # Red
            'HIGH': 'FFDDAA',      # Orange
            'MEDIUM': 'FFFFCC',    # Yellow
            'LOW': 'CCFFCC',       # Green
            'NONE': 'CCFFCC',      # Green
        }.get(risk, 'FFFFFF')
        ws[f'C{row}'].fill = PatternFill(start_color=risk_color, end_color=risk_color, fill_type="solid")

        row += 1

    # Column sizing
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 25

    # Sheet: Risk Calculation Formula
    print("🔢 Creating Risk Calculation sheet...")
    ws = wb.create_sheet("Risk Calculation Formula", 4)

    ws['A1'] = "RISK SCORE CALCULATION METHODOLOGY"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:D1')

    row = 3
    ws[f'A{row}'] = "BASE FORMULA"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    formula_lines = [
        "Risk = (Self-Funding % × 0.40) +",
        "        (Coordinated Score × 0.30) +",
        "        (Unknown Funder % × 0.20) +",
        "        (Automation Score × 0.10)",
    ]
    for line in formula_lines:
        ws[f'A{row}'] = line
        ws[f'A{row}'].font = Font(family='Courier New', size=11)
        row += 1

    row += 1
    ws[f'A{row}'] = "ADJUSTMENTS"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    adjustments = [
        ("CEX Backing", "-0.20", "Institutional backing reduces risk"),
        ("INFRA Automation", "+0.10", "Automation bots increase risk"),
        ("Clean Pattern", "0.10 base", "Minimum for truly clean patterns"),
    ]

    for adj, value, desc in adjustments:
        ws[f'A{row}'] = adj
        ws[f'B{row}'] = value
        ws[f'C{row}'] = desc
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 1
    ws[f'A{row}'] = "RISK THRESHOLDS"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    thresholds = [
        ("CRITICAL", "> 0.30", "🔴 Immediate investigation"),
        ("HIGH", "0.15 - 0.30", "🟠 Monitor closely"),
        ("MEDIUM", "0.05 - 0.15", "🟡 Watch for changes"),
        ("LOW", "< 0.05", "🟢 Normal monitoring"),
    ]

    for tier, range_val, action in thresholds:
        ws[f'A{row}'] = tier
        ws[f'B{row}'] = range_val
        ws[f'C{row}'] = action
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 2
    ws[f'A{row}'] = "COMPONENT WEIGHTS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    weights = [
        ("Self-Funding Percentage", "40%", "Strongest indicator of manipulation"),
        ("Coordination Score", "30%", "Network effect strength"),
        ("Unknown Funder Percentage", "20%", "Unverified sources"),
        ("Automation Score", "10%", "Bot activity level"),
    ]

    for component, weight, reason in weights:
        ws[f'A{row}'] = component
        ws[f'B{row}'] = weight
        ws[f'C{row}'] = reason
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 20

    # Save workbook
    wb.save("FLEX_Comprehensive_Network_Documentation.xlsx")
    print("✅ Enhanced Excel with roles, thresholds, and findings saved")

if __name__ == "__main__":
    add_roles_findings_sheets()
